"""Exporter registry. Each exporter fn(qs) → HttpResponse."""
import csv
import io
import json

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone

_registry: dict = {}


def register(fmt: str):
    def decorator(fn):
        _registry[fmt] = fn
        return fn
    return decorator


def get(fmt: str):
    return _registry.get(fmt)


def available() -> list[str]:
    return list(_registry.keys())


def _filename(fmt: str) -> str:
    ts = timezone.now().strftime('%Y%m%d_%H%M')
    return f'scd_report_{ts}.{fmt}'


def _base_qs(qs):
    return qs.select_related('author', 'project', 'category').prefetch_related('tags')


def _rows(qs):
    for item in _base_qs(qs):
        yield {
            'id':          item.pk,
            'author':      item.author.email,
            'title':       item.title,
            'project':     item.project.name,
            'category':    item.category.name,
            'period_kind': item.get_period_kind_display(),
            'period_start': item.period_start.isoformat(),
            'period_end':   item.period_end.isoformat(),
            'tags':        ','.join(sorted(t.name for t in item.tags.all())),
            'private':     'yes' if item.is_private else 'no',
            'description': item.description,
        }


HEADERS = ['id', 'author', 'title', 'project', 'category',
           'period_kind', 'period_start', 'period_end', 'tags', 'private', 'description']


# ── Text ──────────────────────────────────────────────────────────────────────

@register('txt')
def export_txt(qs) -> HttpResponse:
    lines = []
    for r in _rows(qs):
        lines.append(
            f"[{r['period_start']} – {r['period_end']}] {r['title']}\n"
            f"  Author: {r['author']}  |  {r['project']} / {r['category']}"
            + (f"  |  tags: {r['tags']}" if r['tags'] else '')
            + (f"  [PRIVATE]" if r['private'] == 'yes' else '') + '\n'
            + (f"  {r['description']}\n" if r['description'] else '')
        )
    body = '\n'.join(lines) or '(no entries matched)\n'
    resp = HttpResponse(body, content_type='text/plain; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{_filename("txt")}"'
    return resp


# ── CSV ───────────────────────────────────────────────────────────────────────

@register('csv')
def export_csv(qs) -> HttpResponse:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=HEADERS)
    writer.writeheader()
    for r in _rows(qs):
        writer.writerow(r)
    resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{_filename("csv")}"'
    return resp


# ── JSON ──────────────────────────────────────────────────────────────────────

@register('json')
def export_json(qs) -> HttpResponse:
    data = list(_rows(qs))
    body = json.dumps(data, ensure_ascii=False, indent=2)
    resp = HttpResponse(body, content_type='application/json; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{_filename("json")}"'
    return resp


# ── XLSX ──────────────────────────────────────────────────────────────────────

@register('xlsx')
def export_xlsx(qs) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SCD Report'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E40AF')
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name.replace('_', ' ').title())
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, r in enumerate(_rows(qs), start=2):
        for col_idx, key in enumerate(HEADERS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=r[key])

    # Auto-width (approximate)
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{_filename("xlsx")}"'
    return resp


# ── PDF ───────────────────────────────────────────────────────────────────────

@register('pdf')
def export_pdf(qs) -> HttpResponse:
    from weasyprint import HTML

    rows = list(_rows(qs))
    html_str = render_to_string('reports/pdf_report.html', {
        'rows':    rows,
        'headers': HEADERS,
        'generated_at': timezone.now(),
    })
    pdf_bytes = HTML(string=html_str).write_pdf()
    resp = HttpResponse(pdf_bytes, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{_filename("pdf")}"'
    return resp
