"""Tests for the reports app: access control, preview, and all five exporters."""
import csv
import io
import json
from datetime import date, timedelta

import openpyxl
import pytest

from django.urls import reverse

from apps.accounts.models import User
from apps.entries.models import WorkItem
from apps.taxonomy.models import Category, EntryType, Project, WorkGroup


def _new_entry(project=None, category=None, lab_priority=None, **kwargs):
    """Create a WorkItem and set its project/category/lab_priority M2M relations."""
    item = WorkItem.objects.create(**kwargs)
    if project is not None:
        item.projects.set(project if isinstance(project, (list, tuple)) else [project])
    if category is not None:
        item.categories.set(category if isinstance(category, (list, tuple)) else [category])
    if lab_priority is not None:
        item.lab_priorities.set(lab_priority if isinstance(lab_priority, (list, tuple)) else [lab_priority])
    return item


@pytest.fixture
def admin_user(db):
    return User.objects.create_user(
        username='admin', email='admin@example.com', password='pass',
        role=User.Role.ADMIN,
    )


@pytest.fixture
def auditor_user(db):
    return User.objects.create_user(
        username='auditor', email='auditor@example.com', password='pass',
        role=User.Role.AUDITOR,
    )


@pytest.fixture
def regular_user(db):
    return User.objects.create_user(
        username='user', email='user@example.com', password='pass',
    )


@pytest.fixture
def project(db):
    return Project.objects.create(name='DUNE', slug='dune')


@pytest.fixture
def category(db):
    return Category.objects.create(name='Scientific', slug='scientific')


@pytest.fixture
def entry(db, regular_user, project, category):
    today = date.today()
    start = today - timedelta(days=today.weekday())
    return _new_entry(
        author=regular_user,
        title='Test entry',
        project=project,
        category=category,
        period_kind='week',
        period_start=start,
        period_end=start + timedelta(days=6),
        description='Some work done.',
    )


# ── Access control ────────────────────────────────────────────────────────────

class TestReportAccess:
    def test_anonymous_redirected(self, client):
        assert client.get(reverse('reports:index')).status_code == 302

    def test_regular_user_forbidden(self, client, regular_user):
        client.force_login(regular_user)
        assert client.get(reverse('reports:index')).status_code == 403

    def test_auditor_can_access(self, client, auditor_user):
        client.force_login(auditor_user)
        assert client.get(reverse('reports:index')).status_code == 200

    def test_admin_can_access(self, client, admin_user):
        client.force_login(admin_user)
        assert client.get(reverse('reports:index')).status_code == 200


# ── Preview ───────────────────────────────────────────────────────────────────

class TestReportPreview:
    def test_empty_filters_returns_all(self, client, admin_user, entry):
        client.force_login(admin_user)
        resp = client.post(reverse('reports:preview'), {})
        assert resp.status_code == 200
        assert b'Test entry' in resp.content

    def test_project_filter(self, client, admin_user, entry, project, category, db):
        other = Project.objects.create(name='CMS', slug='cms')
        client.force_login(admin_user)
        resp = client.post(reverse('reports:preview'), {'projects': [other.pk]})
        assert b'Test entry' not in resp.content

    def test_projects_filter_matches_any(self, client, admin_user, entry, project, category, db):
        p2 = Project.objects.create(name='CMS', slug='cms')
        today = date.today()
        start = today - timedelta(days=today.weekday())
        _new_entry(
            project=p2, category=category, author=entry.author, title='Second entry',
            period_kind='week', period_start=start, period_end=start + timedelta(days=6),
            description='x',
        )
        client.force_login(admin_user)
        # Selecting both projects returns entries matching EITHER (OR semantics).
        resp = client.post(reverse('reports:preview'), {'projects': [project.pk, p2.pk]})
        body = resp.content.decode()
        assert 'Test entry' in body
        assert 'Second entry' in body
        # Selecting only the second project excludes the first entry.
        resp = client.post(reverse('reports:preview'), {'projects': [p2.pk]})
        body = resp.content.decode()
        assert 'Test entry' not in body
        assert 'Second entry' in body

    def test_author_email_filter(self, client, admin_user, entry):
        client.force_login(admin_user)
        resp = client.post(reverse('reports:preview'), {'author_email': 'nomatch@x.com'})
        assert b'Test entry' not in resp.content

    def test_entry_type_filter(self, client, admin_user, entry, db):
        weekly = EntryType.objects.create(name='Weekly Report', slug='weekly-report')
        milestone = EntryType.objects.create(name='Milestone', slug='milestone')
        entry.entry_type = weekly
        entry.save(update_fields=['entry_type'])
        client.force_login(admin_user)
        # Matching type shows the entry…
        resp = client.post(reverse('reports:preview'), {'entry_type': weekly.pk})
        assert b'Test entry' in resp.content
        # …a different type filters it out.
        resp = client.post(reverse('reports:preview'), {'entry_type': milestone.pk})
        assert b'Test entry' not in resp.content

    def test_no_match_shows_empty_message(self, client, admin_user):
        client.force_login(admin_user)
        resp = client.post(reverse('reports:preview'), {'author_email': 'nobody@nowhere.com'})
        assert b'No entries matched' in resp.content


# ── Exporters ─────────────────────────────────────────────────────────────────

class TestExporters:
    def _download(self, client, fmt, extra_data=None):
        data = extra_data or {}
        return client.post(reverse('reports:download', kwargs={'fmt': fmt}), data)

    def test_txt_download(self, client, admin_user, entry):
        client.force_login(admin_user)
        resp = self._download(client, 'txt')
        assert resp.status_code == 200
        assert resp['Content-Type'].startswith('text/plain')
        assert b'Test entry' in resp.content

    def test_csv_download(self, client, admin_user, entry):
        client.force_login(admin_user)
        resp = self._download(client, 'csv')
        assert resp.status_code == 200
        assert resp['Content-Type'].startswith('text/csv')
        assert b'Test entry' in resp.content
        assert b'title' in resp.content  # header row

    def test_json_download(self, client, admin_user, entry):
        client.force_login(admin_user)
        resp = self._download(client, 'json')
        assert resp.status_code == 200
        data = json.loads(resp.content)
        assert isinstance(data, list)
        assert any(r['title'] == 'Test entry' for r in data)

    def test_xlsx_download(self, client, admin_user, entry):
        client.force_login(admin_user)
        resp = self._download(client, 'xlsx')
        assert resp.status_code == 200
        assert 'spreadsheetml' in resp['Content-Type']
        # XLSX magic bytes: PK header
        assert resp.content[:2] == b'PK'

    def test_csv_escapes_spreadsheet_formulas(self, client, admin_user, entry):
        entry.title = '=IMPORTXML("https://example.com")'
        entry.description = '@SUM(1,1)'
        entry.save()
        client.force_login(admin_user)

        resp = self._download(client, 'csv')

        rows = list(csv.DictReader(io.StringIO(resp.content.decode())))
        assert rows[0]['title'] == '\'=IMPORTXML("https://example.com")'
        assert rows[0]['description'] == "'@SUM(1,1)"

    def test_xlsx_escapes_spreadsheet_formulas(self, client, admin_user, entry):
        entry.title = '=HYPERLINK("https://example.com")'
        entry.description = '+SUM(1,1)'
        entry.save()
        client.force_login(admin_user)

        resp = self._download(client, 'xlsx')

        wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=False)
        ws = wb.active
        assert ws['C2'].value == '\'=HYPERLINK("https://example.com")'
        assert ws['O2'].value == "'+SUM(1,1)"

    def test_non_spreadsheet_formats_preserve_text_content(self, client, admin_user, entry):
        entry.title = '=Plain text title'
        entry.description = '@Plain text description'
        entry.save()
        client.force_login(admin_user)

        txt_resp = self._download(client, 'txt')
        json_resp = self._download(client, 'json')
        pdf_resp = self._download(client, 'pdf')

        assert b'=Plain text title' in txt_resp.content
        data = json.loads(json_resp.content)
        assert data[0]['title'] == '=Plain text title'
        assert data[0]['description'] == '@Plain text description'
        assert pdf_resp.status_code == 200
        assert pdf_resp.content[:4] == b'%PDF'

    def test_unknown_format_returns_400(self, client, admin_user):
        client.force_login(admin_user)
        resp = self._download(client, 'docx')
        assert resp.status_code == 400

    def test_download_respects_filter(self, client, admin_user, entry):
        client.force_login(admin_user)
        resp = self._download(client, 'json', {'author_email': 'nobody@nowhere.com'})
        data = json.loads(resp.content)
        assert data == []


# ── Group scope filtering (issue #10) ─────────────────────────────────────────

class TestGroupScopeFiltering:
    """
    Group-scoped reports should use WorkItem.group as the authoritative dimension.
    Fallback to author.group only when the entry has no explicit group set.
    """

    def _make_entry(self, author, project, category, group=None):
        today = date.today()
        start = today - timedelta(days=today.weekday())
        return _new_entry(
            author=author,
            title=f'Entry by {author.email}',
            project=project,
            category=category,
            period_kind='week',
            period_start=start,
            period_end=start + timedelta(days=6),
            description='desc',
            group=group,
        )

    def test_entry_assigned_to_leaders_group_is_visible(self, client, db, project, category):
        group_a = WorkGroup.objects.create(name='Group A', slug='group-a')
        group_b = WorkGroup.objects.create(name='Group B', slug='group-b')

        leader = User.objects.create_user(username='leader', email='leader@x.com', password='p',
                                          role=User.Role.GROUP_LEADER, group=group_a)
        outsider = User.objects.create_user(username='out', email='out@x.com', password='p',
                                            group=group_b)
        # Entry is explicitly assigned to group_a even though the author is in group_b
        entry = self._make_entry(outsider, project, category, group=group_a)

        client.force_login(leader)
        resp = client.post(reverse('reports:preview'), {})
        assert resp.status_code == 200
        assert entry.title.encode() in resp.content

    def test_entry_assigned_to_other_group_is_hidden(self, client, db, project, category):
        group_a = WorkGroup.objects.create(name='Group A', slug='group-a')
        group_b = WorkGroup.objects.create(name='Group B', slug='group-b')

        leader = User.objects.create_user(username='leader2', email='leader2@x.com', password='p',
                                          role=User.Role.GROUP_LEADER, group=group_a)
        member = User.objects.create_user(username='mem', email='mem@x.com', password='p',
                                          group=group_a)
        # Entry is explicitly assigned to group_b — should NOT appear for leader of group_a
        entry = self._make_entry(member, project, category, group=group_b)

        client.force_login(leader)
        resp = client.post(reverse('reports:preview'), {})
        assert resp.status_code == 200
        assert entry.title.encode() not in resp.content

    def test_ungrouped_entry_visible_by_author_group(self, client, db, project, category):
        group_a = WorkGroup.objects.create(name='Group A', slug='group-a')

        leader = User.objects.create_user(username='leader3', email='leader3@x.com', password='p',
                                          role=User.Role.GROUP_LEADER, group=group_a)
        member = User.objects.create_user(username='mem2', email='mem2@x.com', password='p',
                                          group=group_a)
        # Entry has no explicit group — falls back to author.group
        entry = self._make_entry(member, project, category, group=None)

        client.force_login(leader)
        resp = client.post(reverse('reports:preview'), {})
        assert resp.status_code == 200
        assert entry.title.encode() in resp.content


class TestNamedTemplateDelete:
    def _make_template(self, user):
        from apps.reports.models import NamedPromptTemplate
        return NamedPromptTemplate.objects.create(
            user=user, name='My Template',
            system_prompt='sys', user_template='tmpl {entries}',
        )

    def test_user_can_delete_own_template(self, client, auditor_user):
        tpl = self._make_template(auditor_user)
        client.force_login(auditor_user)
        resp = client.post(reverse('reports:prompt-template-delete', kwargs={'pk': tpl.pk}))
        assert resp.status_code == 302
        from apps.reports.models import NamedPromptTemplate
        assert not NamedPromptTemplate.objects.filter(pk=tpl.pk).exists()

    def test_user_cannot_delete_others_template(self, client, auditor_user, admin_user):
        tpl = self._make_template(admin_user)
        client.force_login(auditor_user)
        resp = client.post(reverse('reports:prompt-template-delete', kwargs={'pk': tpl.pk}))
        assert resp.status_code == 404

    def test_admin_page_lists_templates(self, client, admin_user, auditor_user):
        self._make_template(auditor_user)
        client.force_login(admin_user)
        resp = client.get(reverse('reports:prompt-templates-admin'))
        assert resp.status_code == 200
        assert b'My Template' in resp.content
        assert b'auditor@example.com' in resp.content

    def test_admin_can_delete_any_template(self, client, admin_user, auditor_user):
        tpl = self._make_template(auditor_user)
        client.force_login(admin_user)
        resp = client.post(reverse('reports:prompt-template-admin-delete', kwargs={'pk': tpl.pk}))
        assert resp.status_code == 302
        from apps.reports.models import NamedPromptTemplate
        assert not NamedPromptTemplate.objects.filter(pk=tpl.pk).exists()

    def test_non_admin_cannot_access_admin_page(self, client, auditor_user):
        client.force_login(auditor_user)
        resp = client.get(reverse('reports:prompt-templates-admin'))
        assert resp.status_code == 403


class TestReportSummary:
    def test_ai_summary_markdown_is_sanitized(self, client, admin_user, entry, monkeypatch):
        def fake_generate(qs, **kwargs):
            return '# Summary\n\n<img src=x onerror=alert(1)> **safe** [bad](javascript:alert(1))'

        monkeypatch.setattr('apps.reports.views.ai_summary.generate', fake_generate)
        client.force_login(admin_user)

        resp = client.post(reverse('reports:summary'), {})

        assert resp.status_code == 200
        assert b'<h1>Summary</h1>' in resp.content
        assert b'<strong>safe</strong>' in resp.content
        assert b'<img' not in resp.content
        assert b'<a>bad</a>' in resp.content
        assert b'<a href="javascript:' not in resp.content
